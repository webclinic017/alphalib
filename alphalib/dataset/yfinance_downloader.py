import os
import time
import traceback
from dataclasses import dataclass
from datetime import datetime
from functools import wraps
from pathlib import Path
from typing import Any, Iterable

import investpy
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from rich import print as rprint
from rich.console import Console
from yfinance import Ticker

from alphalib.data_sources import get_stocks
from alphalib.utils import get_project_root
from alphalib.utils.convertutils import join_dicts


class Downloader:
    def __init__(
        self,
        continue_last_download: bool = True,
        file_prefix: str = "alphalib_",
        sheet_name: str = "",
        start_pos: int = 0,
        primary_col: str = "symbol",
        throttle: int = 2,
        batch_save_size: int = 100,
    ):
        self.continue_last_download = continue_last_download
        self.sheet_name = sheet_name
        self.primary_col = primary_col
        self.throttle = throttle
        self.batch_save_size = batch_save_size
        self.start_pos = start_pos
        self.file_name = str(
            get_project_root()
            .absolute()
            .joinpath("".join([file_prefix, ".xlsx"]))
            .resolve()
        )

    def get_stocks(self) -> pd.DataFrame:
        """Retrieve all stocks."""
        return get_stocks()

    def append_df_to_excel(
        self,
        df: pd.DataFrame,
        startrow: int | None = None,
        **to_excel_kwargs,
    ):
        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(self.file_name):
            df.to_excel(
                self.file_name,
                sheet_name=self.sheet_name,
                startrow=startrow if startrow is not None else 0,
                header=True,
                index=False,
                **to_excel_kwargs,
            )
            return

        # ignore [engine] parameter if it was passed
        if "engine" in to_excel_kwargs:
            to_excel_kwargs.pop("engine")

        # try to open an existing workbook
        work_book = load_workbook(self.file_name, read_only=True, keep_vba=False)  # type: ignore

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and self.sheet_name in work_book.sheetnames:  # type: ignore
            startrow = work_book[self.sheet_name].max_row  # type: ignore

        if startrow is None:
            startrow = 0

        header = False
        if startrow == 0:
            header = True

        # write out the new sheet
        with pd.ExcelWriter(self.file_name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:  # type: ignore
            df.to_excel(
                writer,
                self.sheet_name,
                startrow=startrow,
                header=header,
                index=False,
                **to_excel_kwargs,
            )

        # close the work_book
        work_book.close()

    def create_missing_cols(self, df, target_cols):
        columns = df.columns.tolist()
        missing_cols = list(set(target_cols) - set(columns))
        if len(missing_cols) > 0:
            df[missing_cols] = None

    def check_last_download(self):
        fld_list = []
        lookup = []
        if not self.continue_last_download:
            # Remove the exising file
            Path(self.file_name).unlink(missing_ok=True)
        else:
            if Path(self.file_name).exists():
                df: pd.DataFrame = pd.read_excel(
                    self.file_name,
                    sheet_name=self.sheet_name,
                    engine="openpyxl",
                )
                fld_list = df.columns.tolist()
                fld_list.sort()
                lookup = df[self.primary_col].unique().tolist()  # type: ignore

        return fld_list, lookup

    def __call__(self, fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            stocks = self.get_stocks()
            skip = False
            total_stocks = len(stocks)
            counter = 0
            console = Console()
            has_error = False
            df_result = pd.DataFrame()
            with console.status("[bold green]Downloading stock..."):
                fld_list, lookup = self.check_last_download()
                for stock in stocks.itertuples(index=False, name="Stock"):
                    try:
                        skip = False
                        has_error = False
                        counter = counter + 1
                        if self.start_pos > 0 and counter < self.start_pos:
                            console.log(f"[blue]Skipping {stock.symbol}")  # type: ignore
                            skip = True
                            continue
                        if self.continue_last_download:
                            if stock.symbol in lookup:  # type: ignore
                                console.log(f"[blue]Skipping {stock.symbol}")  # type: ignore
                                skip = True
                                continue

                        ticker: Ticker = yf.Ticker(stock.symbol)  # type: ignore
                        history: pd.DataFrame = ticker.history()
                        if history.empty:
                            console.log(f"[blue]Stock not found. Skipping {stock.symbol}")  # type: ignore
                            # skip = True
                            continue

                        result = fn(
                            *args,
                            ticker=ticker,
                            fld_list=fld_list,
                            stock=stock,
                            **kwargs,
                        )

                        if len(result) > 0 and len(result.columns) > 5:
                            if len(fld_list) == 0:
                                fld_list.extend(result.columns.tolist())
                                fld_list.sort()
                            self.create_missing_cols(result, fld_list)
                            df_result = pd.concat(
                                [df_result, result[fld_list]], ignore_index=True
                            )
                        if len(df_result) >= self.batch_save_size:
                            console.log("[green]Saving fetched stocks...")
                            self.append_df_to_excel(df_result)
                            df_result = pd.DataFrame()

                    except Exception as e:
                        has_error = True
                        rprint(
                            f"Unable to download data for {stock.symbol}-{stock.short_name}",
                            e,
                        )
                        traceback.print_exc()
                        # return # continue
                    finally:
                        if not skip:
                            if not has_error:
                                console.log(
                                    f"[green]{counter}/{total_stocks} - Finish fetching data[/green] {stock.symbol}-{stock.short_name}"
                                )
                            else:
                                console.log(
                                    f"[red]{counter}/{total_stocks} - Error fetching data[/red] {stock.symbol}-{stock.short_name}"
                                )

                            if self.throttle > 0:
                                time.sleep(self.throttle)

            if len(df_result) > 0:
                console.log("[green]Saving remaining fetched stocks...")
                self.append_df_to_excel(df_result)
                df_result = pd.DataFrame()

        return wrapper


@dataclass
class Dataset:
    """Dataset downloader."""

    def __post_init__(self):
        pass

    def __del__(self):
        pass

    def set_stock_info(self, result, stock):
        result["name"] = stock.short_name
        result["symbol"] = stock.symbol
        result["sector"] = stock.sector
        return result

    # def get_stats(self, stats, result, stats_type):
    #     if stats[stats_type]:  # type: ignore
    #         v = stats[stats_type]
    #         if type(v) is dict:
    #             result = {**result, **v}
    #         return result

    @Downloader(file_prefix="stock_info", sheet_name="stock_info")
    def stock_info(self, *_, **kwargs):
        # stock: Iterable[tuple[Any, ...]] = kwargs["stock"]
        ticker: Ticker = kwargs["ticker"]

        # Get stock data
        stock_info = pd.DataFrame([ticker.info])
        return stock_info

    @Downloader(file_prefix="stock_stats", sheet_name="stock_stats")
    def stock_stats(self, *_, **kwargs):
        stock: Iterable[tuple[Any, ...]] = kwargs["stock"]
        ticker: Ticker = kwargs["ticker"]

        stats = ticker.stats()  # type: ignore
        result: dict = {}
        result = join_dicts(result, stats, "defaultKeyStatistics")
        result = join_dicts(result, stats, "financialData")
        result = join_dicts(result, stats, "summaryDetail")
        if not result:
            return pd.DataFrame()
        stock_stats = pd.DataFrame([result])
        stock_stats = self.set_stock_info(stock_stats, stock)
        return stock_stats

    @Downloader(file_prefix="stock_dividends", sheet_name="stock_dividends")
    def stock_dividends(self, *_, **kwargs):
        stock: Iterable[tuple[Any, ...]] = kwargs["stock"]

        # From investpy
        stock_dividends = investpy.get_stock_dividends(stock.symbol, stock.country.lower())  # type: ignore
        if len(stock_dividends) > 0:
            last_10_years = datetime.now().year - 10
            stock_dividends = self.set_stock_info(stock_dividends, stock)
            return stock_dividends[
                pd.DatetimeIndex(stock_dividends["Date"]).year > last_10_years  # type: ignore
            ]
        return pd.DataFrame()

    # @Downloader(file_prefix="alphalib_financials_", sheet_name="stock_financials")
    # def stock_financials(self, *_, **kwargs):
    #     stock: Iterable[tuple[Any, ...]] = kwargs["stock"]
    #     ticker: Ticker = kwargs["ticker"]

    #     stock_financials = ticker.financials
    #     if len(stock_financials) > 0:
    #         stock_financials = stock_financials.T  # type: ignore
    #         stock_financials = self.set_stock_info(stock_financials, stock)
    #         stock_financials.index.name = "Date"
    #         stock_financials.reset_index(inplace=True)
    #         return stock_financials
    #     return pd.DataFrame()

    # stock_cashflow = ticker.cashflow
    # stock_earnings = ticker.earnings
    # stock_balance_sheet = ticker.balance_sheet
    # stock_calendar = ticker.calendar
    # stock_earnings_date  = ticker.earnings_dates
    # stock_recommendations = ticker.recommendations
    # stock_news = ticker.news
    # stock_history = ticker.history()
    # stock_splits = ticker.splits
    # stock_earnings_history = ticker.earnings_history
    # stock_actions = ticker.actions
    # stock_analysis = ticker.analysis
    # stock_stats = ticker.stats()
    # stock_sustainability = ticker.sustainability